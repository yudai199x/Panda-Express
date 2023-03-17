import openpyxl as px
import pandas as pd

mmyyyyy = '2023-01'
purpose = ['教育', 'その他']
product = [['T6XH8', 'T6XN6', 'T6XN5', 'T6XM9', 'T6XP6'], ['T6XZ0']]
dvision = ['その他', 'BiCS4.5以前', 'BiCS5', 'BiCS6']
affiliation = ['解析2', '信技', 'PE技', '品証']

wb = px.load_workbook('doc.xlsx')
ws = wb[mmyyyyy]

d1 = [0 for i in range(4)]
d2 = [0 for i in range(4)]
d = [[0 for i in range(4)] for j in range(4)]
i = 9

while not ws.cell(row=i, column=13).value is None:
    if ws.cell(row=i, column=13).value in purpose:
        x = 0
    else:
        if ws.cell(row=i, column=14).value not in sum(product, []):
            x = 1
        else:
            for j in range(2):
                if ws.cell(row=i, column=14).value in product[j]:
                    x = j + 2
                else:
                    pass 
    
    dt1 = ws.cell(row=i, column=7).value - ws.cell(row=i, column=4).value
    dt2 = ws.cell(row=i, column=9).value - ws.cell(row=i, column=6).value
    dt3 = 0 if dt1 > 0 else 24
    dt4 = ((dt1 + dt3) * 60) + dt2
    
    for j in range(4):
        if ws.cell(row=i, column=24).value == affiliation[j]:
            y = j
        else:
            pass

    d[x][y] += dt4
    i += 1

for x in range(4):
    for y in range(4):
        d1[x] += d[y][x]
        d2[x] += d[x][y]

d3 = [d1[0], d1[2], d1[3]]
d4 = [0 for i in range(4)]

for x in range(4):
    d4[x] += d2[x] - d[x][1]

d5 = d1 + d2
d6 = affiliation + dvision
d7 = pd.Series(d5, index=d6)
d8 = d7 / 60
d9 = d8.to_frame(name=mmyyyyy)

d10 = d3 + d4
d11 = ['解析2', 'PE技', '品証'] + dvision
d12 = pd.Series(d10, index=d11)
d13 = d12 / 60
d14 = d13.to_frame(name=mmyyyyy)

d15 = pd.concat([d9, d14], axis=0)
d15.to_excel('out.xlsx', index=True, header=True)
